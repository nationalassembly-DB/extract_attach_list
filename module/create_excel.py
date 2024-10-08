"""
데이터를 전송 받아서 엑셀에 입력합니다
"""

import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill


from module.create_log import logging


def create_excel(wb, excel_list, excel_path):
    """받은 데이터를 토대로 엑셀을 생성합니다"""
    ws = wb.active

    for item in excel_list:
        last_row = ws.max_row + 1
        ws.cell(row=last_row, column=1, value=item['cmt'])
        ws.cell(row=last_row, column=2, value=item['org'])
        ws.cell(row=last_row, column=7, value=item['name'])
        ws.cell(row=last_row, column=8, value=item['question'])
        ws.cell(row=last_row, column=9, value=item['realfile_name'])
        ws.cell(row=last_row, column=10, value=item['real_path'])
        ws.cell(row=last_row, column=11, value=item['file_name'])

    wb.save(excel_path)


def _has_header(wb, path):
    """엑셀 header가 존재하는지 확인합니다. 존재하지 않을 경우 새로 생성합니다"""
    ws = wb.active
    first_row = ws[1]
    header_exists = any(cell.value for cell in first_row)

    if not header_exists:
        headers = ['위원회', '피감기관', 'BOOK_ID', 'SEQNO', 'FILE_NAME',
                   '국정감사 파일명', '위원', '질의', 'REALFILE_NAME', '실제 경로', '파일명']

        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)
        fill_color = PatternFill(start_color='4f81bd',
                                 end_color='4f81bd', fill_type='solid')

        for col in range(1, 11):
            ws.cell(row=1, column=col).fill = fill_color

    wb.save(path)

    return wb


def _load_excel(excel_path):
    """엑셀을 불러옵니다. 파일이 없는 경우 새로 생성됩니다"""
    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
    else:
        try:
            wb = Workbook()
            wb.save(excel_path)
            wb = load_workbook(excel_path)
        except Exception as e:  # pylint: disable=W0703
            e = "엑셀 파일 생성 오류"
            logging(e, '', os.path.dirname(excel_path))

    return _has_header(wb, excel_path)
